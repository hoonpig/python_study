from openpyxl import load_workbook

# wb = load_workbook("sample_formula.xlsx")
# ws = wb.active


# # 수식그대로 가져온다
# for row in  ws.values:
#     for cell in row:
#         print(cell)

wb = load_workbook("sample_formula.xlsx", data_only=True)   # 수식이 아닌 실제 데이터를 가지고옴
ws = wb.active


# 수식이 아닌 실제 데이터를 가져온다
# evaluate 되지않은 실제 데이터를 가져온다
for row in  ws.values:
    for cell in row:
        print(cell)
