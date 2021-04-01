from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet()      # 새로운 sheet 를 기본이름으로 생성
ws.title = "MySheet"        # sheet 이름 변경
ws.sheet_properties.tabColor = "ff66ff"   # RGB 로 탭 색상 색상설정

ws1 = wb.create_sheet("Your Sheet")         # 주어진 이름으로 Sheet 생성
ws2 = wb.create_sheet("New Sheet" , 2)      # 2번째 시트에 시트생성

new_ws = wb["New Sheet"]    # dict 형태로 sheet 에 접근 가능

print(wb.sheetnames)    # 모든 sheet 이름 확인

# sheet 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copid Sheet"


wb.save("sample.xlsx")