from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "NadoSheet"

ws["A1"] = 1    # A1 이라는 셀에 1이라는 값을 입력
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"])         # A1 셀의 정보를 출력
print(ws["A1"].value)   # A1 셀의 값을 출력
print(ws["A10"].value)  # 값이 없을떈 None 을 출력

print(ws.cell(row=1, column=1).value)   # print(ws["A1"].value) 와 똑같다
print(ws.cell(row=1, column=2).value)
print(ws.cell(column=2, row=1).value)

c = ws.cell(column=3, row=1, value=10)  # ws["c1"].value = 10   / ws["c1"] = 10  와 동일한 뜻
print(c.value)

from random import *

# 반복문을 이용해서 랜덤숫자를 채우기
index = 1
for x in range(1,11):
    for y in range(1,11):
        #ws.cell(row=x, column=y, value = randint(0,100))
        ws.cell(row=x, column=y, value = index)
        index += 1

wb.save("sample.xlsx")