# 과목점수 비중
# - 출석 10
# - 퀴즈1 10
# - 퀴즈2 10
# - 중간 20
# - 기말 30
# - 프로젝트 20
# ------------------------
# 총합 : 100

# 퀴즈2번문제는 만점처리
# H 열에 총점 = sum 이용 , i 열에 성적정보 추가
# : 총점 90이상 A , 80 이상 B, 70 이상 C , 나머지 D
# 출석이 5 미만이면 총점상관없이 F
# 파일명 :  scores.xlsx

from openpyxl import Workbook,load_workbook
from random import *

wb = Workbook()
ws = wb.active

ws.append(["번호","출석","퀴즈1","퀴즈2","중간","기말","프로젝트","총점","성적정보"])
# 내가 한 방법
# ws.append([1,10,8,5,14,26,12])
# ws.append([2,7,3,7,15,24,18])
# ws.append([3,9,5,8,8,12,4])
# ws.append([4,7,8,7,17,21,18])
# ws.append([5,7,8,7,16,25,15])
# ws.append([6,3,5,8,8,17,0])
# ws.append([7,4,9,10,16,27,18])
# ws.append([8,6,6,6,15,19,17])
# ws.append([9,10,10,9,19,30,19])
# ws.append([10,9,8,8,20,25,20])

scores = [
(1,10,8,5,14,26,12),
(2,7,3,7,15,24,18),
(3,9,5,8,8,12,4),
(4,7,8,7,17,21,18),
(5,7,8,7,16,25,15),
(6,3,5,8,8,17,0),
(7,4,9,10,16,27,18),
(8,6,6,6,15,19,17),
(9,10,10,9,19,30,19),
(10,9,8,8,20,25,20)
]

for s in scores:
    ws.append(s)

#for i in range(1,11):
#    ws.append([i , randint(4,10), randint(0,10), 10, randint(0,10), randint(0,30), randint(0,20), randint(0,20) ])

# 퀴즈 2 를 10 점 으로 업데이트
for idx, cell in enumerate(ws["D"]):
    if idx == 0 :
        continue
    cell.value = 10

col_H = ws["H"]
print(f"ws.max_column : {ws.max_column} , ws.max_row : {ws.max_row}")

for idx, score in enumerate(scores, start=2):
    sum_value = sum(score[1:]) - score[3] + 10
    ws.cell(row=idx, column = 8).value = f"=SUM(B{idx}:G{idx})"

    grade = None
    
    if sum_value >= 90:
        grade = "A"
    elif sum_value >= 80:
        grade = "B"
    elif sum_value >= 70:    
        grade = "C"
    else:
        grade = "D"
    
    if score[1] < 5:
        grade = "F"
    
    ws.cell(row=idx, column=9).value = grade


#for row_value in range(2,ws.max_row+1):
#    ws_H_tmp      = "H"+str(row_value)
 
# 가로열 가공
    
#    B_tmp = "B"+str(row_value)
#    C_tmp = "C"+str(row_value)
#    D_tmp = "D"+str(row_value)
#    E_tmp = "E"+str(row_value)
#    F_tmp = "F"+str(row_value)
#    G_tmp = "G"+str(row_value)
#    I_tmp = "I"+str(row_value)
#    ws[ws_H_tmp]  = f"=SUM({B_tmp}:{G_tmp})"
#
#    ws[D_tmp].value=10#

#    sum_tmp = ws[B_tmp].value+ws[C_tmp].value+ws[D_tmp].value+ws[E_tmp].value + ws[F_tmp].value + ws[G_tmp].value
#    print(sum_tmp)
#    if ws[B_tmp].value >= 5 :
#        if sum_tmp >= 90:
#            ws[I_tmp].value = "A"
#        elif sum_tmp >= 80:
#            ws[I_tmp].value = "B"
#        elif sum_tmp >= 70:    
#            ws[I_tmp].value = "C"
#        else:
#            ws[I_tmp].value = "D"
#    else:
#           ws[I_tmp].value = "F"



wb.save("score.xlsx")
