from openpyxl import load_workbook

wb = load_workbook("sample.xlsx")
ws = wb.active


# 번호, 영어, 수학
# 번호, 국어, 영어, 수학
# rows,cols 예시
# 양수일경우
# rows : 아래 / cols : 왼쪽
# 음수일경우
# rows : 위  / cols : 오른쪽

# ws.move_range("B1:C11", rows=0, cols=1)
# ws["B1"].value="국어"

ws.move_range("C1:C11", rows=5, cols=-1)

wb.save("sample_math.xlsx")